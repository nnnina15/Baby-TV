
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "EP01 分鏡表"

COLOR_HEADER     = "2C3E50"
COLOR_OPENING    = "FAD7A0"
COLOR_CHALLENGE1 = "A9DFBF"
COLOR_CHALLENGE2 = "AED6F1"
COLOR_CHALLENGE3 = "D2B4DE"
COLOR_LEARNING   = "F9E79F"
COLOR_SONG       = "FADBD8"
COLOR_ENDING     = "D5DBDB"
COLOR_INTERACT   = "F0E6FF"

# ── Character reference descriptions ──
MOCHI  = "Mochi (small fluffy cream-colored chibi bear, soft furry texture, pink nose, round ears with pink inner, large brown sparkly eyes, rosy cheeks, yellow star patch on belly)"
BLOOP  = "Bloop (teal mint water-drop frog, shiny smooth plastic body, round head with pointed top, dark teal polka dots, large teal eyes, small red smiling mouth, webbed sticky fingers and toes)"
SUNNY  = "Sunny (small fluffy bright yellow baby chick, soft furry round body, orange beak, large round brown eyes, spiky yellow hair tuft, small orange feet)"
ZIGGY  = "Ziggy (rainbow caterpillar, round red shiny plastic head, red ball-tipped antennae, large white eyes, happy smile, colorful stacked sphere body: red orange yellow green blue purple)"
PUDDING = "Pudding (small fluffy lavender purple elephant, soft furry texture, large round floppy ears, small curved trunk, sleepy half-closed violet eyes, tufted hair on head)"
STYLE  = "3D render, toy figurine style, soft toy aesthetic, pastel background, studio lighting, children's TV show, ultra clean"

def hfill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

thin = Side(style="thin", color="AAAAAA")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap_align = Alignment(wrap_text=True, vertical="top")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
body_font = Font(name="Arial", size=9)

col_widths = {"A":4,"B":10,"C":18,"D":14,"E":38,"F":20,"G":9,"H":65}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

headers = ["#","時間","段落","鏡頭類型","畫面描述","情緒關鍵字","建議秒數","AI 生成 Prompt"]
for ci, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font = header_font
    cell.fill = hfill(COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
ws.row_dimensions[1].height = 28

scenes = [
    (COLOR_OPENING, 1, "0:00–0:08", "Opening 片頭",
     "WS 全景",
     "Wobbly World 村莊早晨全景：粉彩色調雲朵飄浮，地面有彈性，彩色小房子點綴其中，陽光從右側灑入",
     "溫暖 / 歡迎 / 魔幻", 8,
     f"Wobbly World village morning wide shot, soft pastel colors, bouncy ground, rainbow clouds, cute round buildings with flowers, warm golden sunrise lighting, no characters, {STYLE}"),

    (COLOR_OPENING, 2, "0:08–0:18", "Opening 角色登場",
     "MS 中景",
     "Mochi 從左方跳入畫面，雙手揮舞，臉上掛著大大的笑容，直視鏡頭",
     "開心 / 活潑 / 友善", 10,
     f"{MOCHI}, bouncing into frame from left, waving both arms enthusiastically, huge happy smile, looking directly at camera, Wobbly World pastel village background, {STYLE}"),

    (COLOR_OPENING, 3, "0:18–0:30", "Opening 神秘物體",
     "CU 特寫",
     "Mochi 雙手高舉一個用綠葉包裹的神秘物體，眼睛睜大充滿好奇，葉片邊緣透出微微金色光暈",
     "神秘 / 好奇 / 期待", 12,
     f"{MOCHI}, holding mysterious object wrapped in bright green leaf with both hands raised high, wide curious eyes, faint golden glow peeking through leaf edges, close-up shot, soft pastel background, {STYLE}"),

    (COLOR_INTERACT, 4, "0:30–0:45", "Opening 🎯 互動 1",
     "MS + 互動框",
     "Mochi 直視鏡頭、微微前傾，螢幕下方出現互動提示框，手指指向觀眾，邀請一起找形狀",
     "互動 / 邀請 / 期待", 15,
     f"{MOCHI}, leaning forward slightly toward camera, pointing one finger directly at viewer, friendly inviting expression, interactive speech bubble text below frame, Wobbly World background, {STYLE}"),

    (COLOR_CHALLENGE1, 5, "0:45–0:55", "挑戰1 跑向池塘",
     "WS 追蹤鏡頭",
     "Mochi 拿著葉包物體小跑向畫面右方，背景切換至荷花池——藍綠色水面、睡蓮葉、垂柳",
     "冒險 / 急迫 / 期待", 10,
     f"{MOCHI}, running with leaf-wrapped object toward right, side-view tracking shot, colorful pond background with lily pads on blue-green water, willow tree, pastel Wobbly World setting, {STYLE}"),

    (COLOR_CHALLENGE1, 6, "0:55–1:10", "挑戰1 Bloop 現身",
     "CU 特寫",
     "Bloop 從水中緩緩探出頭，眼睛半閉，表情慵懶溫柔，水面泛起小漣漪",
     "慵懶 / 溫柔 / 緩慢", 15,
     f"{BLOOP}, slowly peeking out from pond surface, half-closed sleepy eyes, gentle lazy expression, small water ripples around face, lily pad nearby, soft blue-green water, {STYLE}"),

    (COLOR_CHALLENGE1, 7, "1:10–1:30", "挑戰1 研究物體",
     "MS 中景",
     "Mochi 和 Bloop 並肩蹲著，一起盯著葉包物體，兩人眉頭微皺充滿思考感，背景是荷花池",
     "困惑 / 思考 / 合作", 20,
     f"{MOCHI} and {BLOOP}, side by side crouching together, both staring at leaf-wrapped mystery object on ground, thinking expressions with slight furrowed brows, pond background with lily pads, {STYLE}"),

    (COLOR_CHALLENGE1, 8, "1:30–1:55", "挑戰1 葉子卡住",
     "CU 特寫",
     "兩人同時用力拉葉片，手部特寫——葉子紋絲不動，兩人臉上出現汗滴符號",
     "努力 / 困難 / 搞笑", 25,
     f"{MOCHI} and {BLOOP}, both pulling hard on green leaf wrapper with all their strength, close-up of hands and faces showing strain effort, leaf not moving, comic sweat drops appearing, funny moment, {STYLE}"),

    (COLOR_INTERACT, 9, "1:55–2:10", "挑戰1 🎯 互動 2",
     "MS + 互動框",
     "Mochi 轉向鏡頭做「拉」的手勢，螢幕出現 'Pull! Pull! Pull!' 文字提示，鼓勵觀眾一起說",
     "互動 / 一起 / 努力", 15,
     f"{MOCHI}, facing camera doing pulling gesture with both arms, interactive colorful text overlay 'Pull! Pull! Pull!' in bubbly font, encouraging wide-open eyes, Wobbly World pond background, {STYLE}"),

    (COLOR_CHALLENGE2, 10, "2:10–2:20", "挑戰2 前往花田",
     "WS 全景",
     "Mochi 和 Bloop 並排走向右方，背景轉換至金黃色向日葵花田，遠處 Sunny 在揮手",
     "開心 / 陽光 / 溫暖", 10,
     f"{MOCHI} and {BLOOP}, walking together toward right, wide shot, bright yellow sunflower field background, {SUNNY} visible waving in distance, warm sunshine atmosphere, {STYLE}"),

    (COLOR_CHALLENGE2, 11, "2:20–2:40", "挑戰2 Sunny 出場",
     "MS 中景",
     "Sunny 全力衝刺跑來迎接，翅膀拍動，眼睛圓又大，整個身體幾乎跳離地面",
     "興奮 / 熱情 / 活潑", 20,
     f"{SUNNY}, running at full speed toward camera with wings flapping rapidly, huge wide sparkling eyes, almost leaping off ground, motion lines behind body, high energy happy expression, sunflower field background, {STYLE}"),

    (COLOR_CHALLENGE2, 12, "2:40–3:15", "挑戰2 感受形狀數數",
     "MS + 文字疊加",
     "三人圍著葉包物體用手觸摸，Sunny 每數一個角，螢幕對應出現 1、2、3、4、5 的跳出數字動畫",
     "專注 / 發現 / 學習", 35,
     f"{MOCHI}, {BLOOP}, and {SUNNY}, all surrounding leaf-wrapped object touching it to feel the shape, counting gestures, animated numbers 1-2-3-4-5 popping up on screen with star burst effects, discovery expressions, educational children moment, {STYLE}"),

    (COLOR_INTERACT, 13, "3:15–3:30", "挑戰2 🎯 互動 3",
     "MS + 互動框",
     "Sunny 面向鏡頭伸出手指逐一張開數 1-5，螢幕數字大大顯示，邀請觀眾跟著數",
     "互動 / 學習 / 開心", 15,
     f"{SUNNY}, facing camera extending wing feathers one by one counting 1 to 5, large colorful animated numbers on screen, cheerful interactive moment, bright yellow energy, sunflower background, {STYLE}"),

    (COLOR_CHALLENGE3, 14, "3:40–3:55", "挑戰3 Ziggy 出場",
     "MS 中景",
     "Ziggy 用身體 S 形扭動進入畫面，懷裡抱著一堆小樹枝，表情充滿點子",
     "創意 / 能量 / 驚喜", 15,
     f"{ZIGGY}, wiggling in S-shape motion into frame from right side, carrying bundle of small sticks, light bulb moment expression, energetic bouncy movement, Wobbly World forest clearing background, {STYLE}"),

    (COLOR_CHALLENGE3, 15, "3:55–4:25", "挑戰3 拼出星形",
     "WS 俯視 → CU",
     "Ziggy 在地面排列五根樹枝，先俯視全景看清楚星形圖案，再推近特寫——圖案閃閃發光",
     "創意 / 驚喜 / 發現", 30,
     f"{ZIGGY}, arranging five sticks on grass ground forming star shape, overhead aerial view showing complete star pattern clearly, then close-up zoom with subtle sparkle glow on star outline, educational discovery moment, {STYLE}"),

    (COLOR_CHALLENGE3, 16, "4:25–4:30", "挑戰3 眾人驚呼",
     "MS 群體反應",
     "Mochi、Bloop、Sunny 三人同時張大嘴巴，眼睛發光，四顆星星從眼睛飛出（經典卡通驚喜反應）",
     "驚喜 / 高潮 / 歡喜", 5,
     f"{MOCHI}, {BLOOP}, and {SUNNY}, all three with mouths wide open in amazement, star sparkles flying out from eyes, classic cartoon surprise reaction, group reaction shot, {STYLE}"),

    (COLOR_CHALLENGE3, 17, "4:30–4:55", "挑戰3 揭開物體",
     "CU → MS 中景",
     "Mochi 緩緩打開葉片——先從縫隙透出耀眼金光（特寫），再拉遠看到金色星星落下，四人仰頭目送",
     "高潮 / 驚喜 / 魔幻", 25,
     f"{MOCHI}, slowly opening green leaf wrapper, golden light streaming through cracks (close-up), then wide pull back revealing glowing golden five-pointed star floating out with sparkle trails, all four characters watching in awe, magical dramatic reveal, {STYLE}"),

    (COLOR_INTERACT, 18, "4:55–5:30", "挑戰3 🎯 互動 4",
     "MS + 互動框",
     "螢幕中央的金色星星閃閃發光，Mochi 手指螢幕邀請觀眾「指向星形」，角落出現手指圖示",
     "互動 / 成就感 / 開心", 35,
     f"Golden glowing five-pointed star shape centered on screen with sparkle animation, {MOCHI} pointing at star with excited expression inviting audience to point, finger pointer icon in corner of screen, interactive children participation moment, {STYLE}"),

    (COLOR_LEARNING, 19, "5:30–5:55", "學習段 星星特寫",
     "ECU 超特寫",
     "金色星星佔滿畫面，五個角依序由暗轉亮（逐一點亮），螢幕右上角出現白色文字標籤「STAR」",
     "清晰 / 重點 / 教育", 25,
     f"Extreme close-up of golden five-pointed star shape filling screen, each of five points lighting up sequentially with glow effect one by one, white bold text label 'STAR' appearing top right corner, educational clarity moment, dark soft background with light effect, {STYLE}"),

    (COLOR_LEARNING, 20, "5:55–6:30", "學習段 Sunny 帶數數",
     "MS + 文字疊加",
     "Sunny 站在金色星星旁，每指一個角就亮起，螢幕大字依序顯示 1→2→3→4→5",
     "學習 / 節奏 / 快樂", 35,
     f"{SUNNY}, standing beside large glowing golden star, pointing wing to each point as it lights up in sequence, large rounded bubbly numbers 1-2-3-4-5 appearing on screen, cheerful educational counting moment, soft pastel background, {STYLE}"),

    (COLOR_LEARNING, 21, "6:30–7:00", "學習段 生活中星形",
     "蒙太奇 快切",
     "快速連切：星形餅乾 → 警徽 → 兒童畫星星 → 星形貼紙 → 星形切模，每個畫面約 3 秒",
     "聯想 / 發現 / 生活感", 30,
     f"Montage of real-world star shapes illustrated in Wobbly World toy style: star-shaped cookie, sheriff badge star, child's star drawing, gold star sticker, star cookie cutter mold, each object rendered as cute 3D toy with soft pastel colors, {STYLE}"),

    (COLOR_INTERACT, 22, "7:00–7:30", "學習段 🎯 互動 5",
     "CU → MS + 互動框",
     "Mochi 先特寫雙手張開五指比出星形，再拉遠看全身示範，螢幕旁出現「Spread your fingers!」",
     "互動 / 身體律動 / 歡樂", 30,
     f"{MOCHI}, close-up of fluffy paws spreading five fingers wide like star shape, then wider body shot showing full star pose, interactive speech bubble 'Spread your fingers!' on screen, encouraging warm smile, {STYLE}"),

    (COLOR_LEARNING, 23, "7:30–7:50", "學習段 Pudding 登場",
     "MS 中景",
     "Pudding 從右側緩緩步入，睡眼惺忪，說出台詞後眾人齊齊抬頭，帶著詩意柔和的驚嘆",
     "溫柔 / 詩意 / 靜謐", 20,
     f"{PUDDING}, slowly shuffling in from right side, characteristic sleepy droopy eyes, gentle dreamy expression, other characters turning heads upward with soft wonder expressions, warm evening golden light beginning, {STYLE}"),

    (COLOR_LEARNING, 24, "7:50–8:30", "學習段 夜空星星",
     "WS 夜景動畫",
     "背景切換至深藍紫色夜空，大小不一的星形輪廓與實心星交替閃爍，角色小小地站在下方仰望",
     "夢幻 / 美麗 / 奇妙", 40,
     f"Beautiful night sky background deep blue purple gradient, twinkling stars various sizes outlined and filled alternating with glow, small silhouettes of all five characters below looking up at sky, dreamy magical atmosphere, {STYLE}"),

    (COLOR_INTERACT, 25, "8:30–9:00", "學習段 🎯 互動 6",
     "MS + 文字疊加",
     "Mochi 面向鏡頭說三次「Star!」，每次說完螢幕都出現發光的「STAR」字樣，搭配星星爆炸小動畫",
     "互動 / 重複學習 / 強化", 30,
     f"{MOCHI}, facing camera with excited open-mouth expression saying Star three times, glowing text 'STAR' appearing with star burst effect each time, word reinforcement learning moment, bright cheerful colors, {STYLE}"),

    (COLOR_SONG, 26, "9:00–9:15", "歌曲段 全員登場",
     "WS 集合鏡頭",
     "五隻角色同時出現，音符從各角色身邊飄出，能量滿溢",
     "歡慶 / 能量 / 完整", 15,
     f"All five Wobbly World characters together in one scene: {MOCHI}, {BLOOP}, {SUNNY}, {ZIGGY}, {PUDDING}, standing in a cheerful group, musical notes floating around them, colorful confetti and sparkles, maximum joy energy, {STYLE}"),

    (COLOR_SONG, 27, "9:15–9:45", "歌曲段 Verse 1 + Chorus",
     "MS 輪播",
     "角色們隨 4/4 拍節奏搖擺，數字 1-5 依歌詞出現，星星的五個角逐一點亮配合節拍",
     "節奏感 / 快樂 / 學習", 30,
     f"All five characters swaying happily to music rhythm, animated numbers 1-2-3-4-5 appearing with lyrics, golden star points lighting up one by one in sync with beat, musical notes floating, singing open-mouth expressions, colorful energetic scene, {STYLE}"),

    (COLOR_SONG, 28, "9:45–10:15", "歌曲段 Verse 2 互動",
     "MS → WS + 互動框",
     "角色們把手臂高舉做出星形，螢幕出現「Reach your arms up!」邀請觀眾跟著做",
     "身體律動 / 互動 / 歡樂", 30,
     f"All five characters reaching arms/wings/trunk up high making star shape poses, full body shot, interactive text 'Reach your arms up!' on screen, audience participation moment, bright energetic colors, confetti, {STYLE}"),

    (COLOR_SONG, 29, "10:15–10:30", "歌曲段 尾聲",
     "WS 高潮",
     "所有角色齊聲「Hip hip hooray!」，彩色碎紙飄落，金色大星星在畫面最上方閃閃發光",
     "高潮 / 成就 / 喜悅", 15,
     f"All five Wobbly World characters celebrating with arms raised, colorful confetti raining down everywhere, large glowing golden star bursting with light at top of frame, peak celebration maximum happiness, {STYLE}"),

    (COLOR_ENDING, 30, "10:30–10:50", "結尾 圍坐回顧",
     "MS 圓形構圖",
     "五隻角色坐成一個溫馨的圓圈，光線轉為暖橘色夕陽調，氛圍從興奮轉向溫暖平靜",
     "溫馨 / 回顧 / 完整感", 20,
     f"All five characters sitting in a cozy circle together, warm orange sunset lighting, calm peaceful relaxed atmosphere, gentle content smiles, Wobbly World meadow background with soft evening glow, {STYLE}"),

    (COLOR_ENDING, 31, "10:50–11:10", "結尾 各角色台詞",
     "CU 逐一 → MS",
     "鏡頭依序給 Bloop、Sunny、Ziggy、Pudding 特寫說台詞，最後拉遠看到五人在一起",
     "回顧 / 成長 / 溫暖", 20,
     f"Sequential close-up portrait shots of {BLOOP}, then {SUNNY}, then {ZIGGY}, then {PUDDING}, each with gentle warm expression speaking, soft bokeh background, then final wide pull back showing all five together, warm wrap-up, {STYLE}"),

    (COLOR_INTERACT, 32, "11:10–11:25", "結尾 最終互動",
     "MS + 文字疊加",
     "全體角色齊聲數 1-2-3-4-5，金色星星從畫面中心爆發飛出，螢幕數字同步出現",
     "最終互動 / 成就感 / 告別", 15,
     f"All five characters counting together 1-2-3-4-5 with matching gestures, golden star burst exploding from screen center, synchronized number text appearing, final triumphant celebratory moment, sparkle effects, {STYLE}"),

    (COLOR_ENDING, 33, "11:25–11:30", "結尾 片尾卡",
     "WS 靜態片尾",
     "Wobbly World Logo 居中，訂閱按鈕在下方有彈跳動畫，五隻角色在兩側揮手道別",
     "友善 / 告別 / 品牌", 5,
     f"Wobbly World logo end card centered, subscribe button with bounce animation below, all five characters {MOCHI} {BLOOP} {SUNNY} {ZIGGY} {PUDDING} on sides waving goodbye, clean brand composition, {STYLE}"),
]

for row_idx, scene in enumerate(scenes, 2):
    bg, num, time, section, shot, visual, emotion, secs, prompt = scene
    values = [num, time, section, shot, visual, emotion, secs, prompt]
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        cell.fill = hfill(bg)
        cell.border = thin_border
        cell.alignment = wrap_align
        if ci == 1:
            cell.alignment = Alignment(horizontal="center", vertical="top")
            cell.font = Font(name="Arial", bold=True, size=9)
        elif ci == 7:
            cell.alignment = Alignment(horizontal="center", vertical="top")
            cell.font = body_font
        else:
            cell.font = body_font
    ws.row_dimensions[row_idx].height = 80

ws.freeze_panes = "A2"

ls = wb.create_sheet("圖例 + 角色說明")
ls.column_dimensions["A"].width = 22
ls.column_dimensions["B"].width = 70
ls["A1"] = "段落顏色圖例"
ls["A1"].font = Font(name="Arial", bold=True, size=12)
legend_items = [
    (COLOR_OPENING,    "Opening 片頭 (0:00–0:45)"),
    (COLOR_CHALLENGE1, "挑戰 1 — Ask Bloop (0:45–2:10)"),
    (COLOR_CHALLENGE2, "挑戰 2 — Ask Sunny (2:10–3:30)"),
    (COLOR_CHALLENGE3, "挑戰 3 — Ask Ziggy (3:40–5:30)"),
    (COLOR_LEARNING,   "學習段 Learning (5:30–9:00)"),
    (COLOR_SONG,       "歌曲段 Song (9:00–10:30)"),
    (COLOR_ENDING,     "結尾 Ending (10:30–11:30)"),
    (COLOR_INTERACT,   "🎯 觀眾互動場景"),
]
tb = Border(left=Side(style="thin",color="AAAAAA"),right=Side(style="thin",color="AAAAAA"),
            top=Side(style="thin",color="AAAAAA"),bottom=Side(style="thin",color="AAAAAA"))
for i, (color, label) in enumerate(legend_items, 3):
    c1 = ls.cell(row=i, column=1, value=""); c1.fill = hfill(color); c1.border = tb
    c2 = ls.cell(row=i, column=2, value=label); c2.font = Font(name="Arial", size=10)
    ls.row_dimensions[i].height = 20

ls["A13"] = "角色 AI Prompt 描述"
ls["A13"].font = Font(name="Arial", bold=True, size=12)
chars = [
    ("Mochi",   MOCHI),
    ("Bloop",   BLOOP),
    ("Sunny",   SUNNY),
    ("Ziggy",   ZIGGY),
    ("Pudding", PUDDING),
    ("共用風格", STYLE),
]
for i, (name, desc) in enumerate(chars, 15):
    ls.cell(row=i, column=1, value=name).font = Font(name="Arial", bold=True, size=10)
    c = ls.cell(row=i, column=2, value=desc)
    c.font = Font(name="Arial", size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ls.row_dimensions[i].height = 40

wb.save(r"N:\Work\git-repositories\Baby-TV\EP01_分鏡表.xlsx")
print("Done")
